from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from inventory.models import GoodsReceipt, InventoryRequest, Issuance, Product, PurchaseOrder, StockTransaction, Supplier
from inventory.services import post_goods_receipt


def parse_date(value):
    if value in (None, ""):
        return date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return date.today()


def parse_decimal(value):
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    cleaned = str(value).replace(",", "").strip()
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0.00")


class Command(BaseCommand):
    help = "Import products and operational history from the provided inventory workbook."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Absolute path to the Excel workbook.")
        parser.add_argument("--replace", action="store_true", help="Delete existing ERP data before import.")

    @transaction.atomic
    def handle(self, *args, **options):
        workbook_path = Path(options["path"])
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        if options["replace"]:
            self.stdout.write("Clearing existing ERP data...")
            StockTransaction.objects.all().delete()
            GoodsReceipt.objects.all().delete()
            Issuance.objects.all().delete()
            InventoryRequest.objects.all().delete()
            PurchaseOrder.objects.all().delete()
            Supplier.objects.all().delete()
            Product.objects.all().delete()

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        product_cache = {}

        def get_product(sku, name="", category="", product_type="", location="", notes=""):
            if not sku:
                return None
            key = str(sku).strip()
            if key in product_cache:
                return product_cache[key]
            product, _ = Product.objects.get_or_create(
                sku=key,
                defaults={
                    "name": str(name or key).strip(),
                    "category": category,
                    "product_type": str(product_type or "").strip(),
                    "default_location": str(location or "").strip(),
                    "notes": str(notes or "").strip(),
                },
            )
            changed = False
            if name and not product.name:
                product.name = str(name).strip()
                changed = True
            if category and not product.category:
                product.category = category
                changed = True
            if product_type and not product.product_type:
                product.product_type = str(product_type).strip()
                changed = True
            if location and not product.default_location:
                product.default_location = str(location).strip()
                changed = True
            if notes and not product.notes:
                product.notes = str(notes).strip()
                changed = True
            if changed:
                product.save()
            product_cache[key] = product
            return product

        for sheet_name, category in [
            ("RWHU_Inventory", "RWHU"),
            ("Stationary_Inventory", "Stationary"),
            ("Progressive_Inventory", "Progressive"),
        ]:
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                sku, name, product_type, _stock, location, _status, notes = (list(row) + [None] * 7)[:7]
                if not sku or not name:
                    continue
                get_product(sku, name=name, category=category, product_type=product_type, location=location, notes=notes)

        if "Vendor_List " in workbook.sheetnames:
            worksheet = workbook["Vendor_List "]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                sku, vendor_name, contact, address, email, notes = (list(row) + [None] * 6)[:6]
                get_product(sku)
                if vendor_name:
                    Supplier.objects.get_or_create(
                        name=str(vendor_name).strip(),
                        defaults={
                            "contact_person": str(contact or "").strip(),
                            "address": str(address or "").strip(),
                            "email": str(email or "").strip(),
                            "notes": str(notes or "").strip(),
                        },
                    )

        if "PO" in workbook.sheetnames:
            worksheet = workbook["PO"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                po_number, issue_date, vendor_name, specs, quantity_ordered, po_amount, status, arrived_by, location, quantity_received, notes = (
                    list(row) + [None] * 11
                )[:11]
                if not po_number:
                    continue
                supplier = None
                if vendor_name:
                    supplier, _ = Supplier.objects.get_or_create(name=str(vendor_name).strip())
                PurchaseOrder.objects.get_or_create(
                    po_number=str(po_number).strip(),
                    defaults={
                        "issue_date": parse_date(issue_date),
                        "source_request": None,
                        "supplier": supplier,
                        "specifications": str(specs or "").strip(),
                        "quantity_ordered": parse_decimal(quantity_ordered),
                        "shortage_quantity": Decimal("0.00"),
                        "po_amount": parse_decimal(po_amount),
                        "status": PurchaseOrder.STATUS_PENDING if not status else str(status).strip().lower(),
                        "system_generated": False,
                        "arrived_by": str(arrived_by or "").strip(),
                        "location": str(location or "").strip(),
                        "quantity_received": parse_decimal(quantity_received),
                        "notes": str(notes or "").strip(),
                    },
                )

        if "GRN" in workbook.sheetnames:
            worksheet = workbook["GRN"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                grn_number, po_number, quantity_received, grn_date, received_by, location, notes = (
                    list(row) + [None] * 7
                )[:7]
                if not grn_number:
                    continue
                purchase_order = PurchaseOrder.objects.filter(po_number=str(po_number).strip()).first() if po_number else None
                receipt, _ = GoodsReceipt.objects.get_or_create(
                    grn_number=str(grn_number).strip(),
                    defaults={
                        "purchase_order": purchase_order,
                        "product": purchase_order.product if purchase_order and purchase_order.product else None,
                        "quantity_received": parse_decimal(quantity_received),
                        "grn_date": parse_date(grn_date),
                        "received_by": str(received_by or "").strip(),
                        "location": str(location or "").strip(),
                        "notes": str(notes or "").strip(),
                        "auto_issued_quantity": Decimal("0.00"),
                    },
                )
                if receipt.product and not receipt.posted:
                    post_goods_receipt(receipt)

        if "Stock_in" in workbook.sheetnames:
            worksheet = workbook["Stock_in"]
            for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                order_date, sku, name, product_type, quantity, vendor_name, location, received_by, notes = (
                    list(row) + [None] * 9
                )[:9]
                if not sku:
                    continue
                product = get_product(sku, name=name, product_type=product_type, location=location)
                if vendor_name:
                    Supplier.objects.get_or_create(name=str(vendor_name).strip())
                StockTransaction.objects.get_or_create(
                    reference_type="Opening Stock",
                    reference_number=f"OPEN-{index}",
                    product=product,
                    defaults={
                        "quantity": parse_decimal(quantity),
                        "transaction_type": StockTransaction.TYPE_RECEIPT,
                        "location": str(location or product.default_location or "").strip(),
                        "transaction_date": parse_date(order_date),
                        "notes": str(notes or received_by or "").strip(),
                    },
                )

        request_status_map = {}
        if "Request_Records" in workbook.sheetnames:
            worksheet = workbook["Request_Records"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                request_number = row[0]
                status = row[9] if len(row) > 9 else None
                if request_number and status:
                    request_status_map[str(request_number).strip().lower()] = str(status).strip().lower()

        if "Requests" in workbook.sheetnames:
            worksheet = workbook["Requests"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                request_number, request_date, requested_by, department, location, sku, name, product_type, quantity, manager_email, _email = (
                    list(row) + [None] * 11
                )[:11]
                if not request_number or not sku:
                    continue
                product = get_product(sku, name=name, product_type=product_type, location=location)
                raw_status = request_status_map.get(str(request_number).strip().lower(), "pending")
                status = {
                    "approved": InventoryRequest.STATUS_APPROVED,
                    "issued": InventoryRequest.STATUS_ISSUED,
                    "pending": InventoryRequest.STATUS_PENDING,
                    "rejected": InventoryRequest.STATUS_REJECTED,
                }.get(raw_status, InventoryRequest.STATUS_PENDING)
                InventoryRequest.objects.get_or_create(
                    request_number=str(request_number).strip(),
                    defaults={
                        "request_date": parse_date(request_date),
                        "requested_by": str(requested_by or "").strip(),
                        "department": str(department or "").strip(),
                        "location": str(location or "").strip(),
                        "product": product,
                        "available_quantity": Decimal("0.00"),
                        "quantity_requested": parse_decimal(quantity),
                        "short_quantity": Decimal("0.00"),
                        "manager_email": str(manager_email or "").strip(),
                        "approval_status": status,
                        "fulfillment_status": InventoryRequest.FULFILLMENT_PENDING,
                    },
                )

        self.stdout.write(self.style.SUCCESS("Workbook import completed successfully."))
